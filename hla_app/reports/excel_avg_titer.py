"""Генерация Excel по титру антител A, B и DRB1.

В файле находятся правила именования итоговых Excel-файлов, подготовка
таблиц по антителам, объединение Class I и Class II и само оформление листа
через `create_avg_titer_excel_from_parsed*`. Все изменения формата этого
отчета и логики его объединения удобно искать здесь.
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from hla_app.data.luminex_parser import ParsedLuminexCsv
from hla_app.reports.excel_common import (
    auto_width,
    center_range,
    make_bold,
    make_center,
    paint_header_row,
    protect_sheet,
)
from hla_app.storage.fs_ops import delete_file
from hla_app.utils.patient_name import (
    format_patient_short_name,
    normalize_patient_name,
)
from hla_app.utils.validators import cap_hyphenated_fio_part, format_ddmmyyyy

# --- Именование итоговых Excel-файлов и ключи группировки ---

_INVALID_FILENAME_CHARS_RE = re.compile(r'[\\/:*?"<>|]')


def build_patient_prefix(patient: str) -> str:
    patient = normalize_patient_name(patient)

    parts = [p for p in patient.split() if p]
    if not parts:
        return "Пациент"

    surname_raw = re.sub(r"[^A-Za-zА-Яа-яЁё-]+", "", parts[0]).strip("-")
    surname = cap_hyphenated_fio_part(surname_raw) if surname_raw else "Пациент"

    rest = "".join(parts[1:])
    letters = re.findall(r"[A-Za-zА-Яа-яЁё]", rest)[:2]
    initials = "".join(letter.upper() for letter in letters)

    prefix = f"{surname}_{initials}" if initials else surname
    return _INVALID_FILENAME_CHARS_RE.sub("_", prefix)


def build_titer_excel_filename(
    patient: str,
    hla_class: str,
    batch_date: str,
    num_register_text: str | None = None,
    *,
    patient_dir_name_text: str | None = None,
) -> str:
    prefix = f"{num_register_text}__" if (num_register_text or "") else ""
    patient_part = patient_dir_name_text or build_patient_prefix(patient)

    return f"{prefix}{patient_part}__AvgTiter_Class_{hla_class}__{batch_date}.xlsx"


def build_titer_group_key(parsed: ParsedLuminexCsv) -> tuple[str, str]:
    # Используем сырые patient и batch_date без нормализации.
    return parsed.patient or "", parsed.batch_date or ""


# --- Преобразование специфичности и сборка таблиц антител ---


def split_specificity_parts(value: str) -> tuple[str, str]:
    specificity = str(value or "").strip()
    specificity = specificity.rsplit(":", 1)[0]

    if "*" not in specificity:
        return specificity, ""

    locus, allele_value = specificity.split("*", 1)
    return locus, allele_value


def build_titer_dataframe(
    *,
    antibodies: list[dict[str, str]],
    hla_class: str,
    min_titer: int = 0,
) -> pd.DataFrame:
    dataframe = pd.DataFrame(antibodies)

    if dataframe.empty or "Specificity" not in dataframe.columns:
        return pd.DataFrame(columns=["Локус", "Значение", "MFI"])

    dataframe["Raw Value"] = pd.to_numeric(
        dataframe.get("Raw Value"),
        errors="coerce",
    ).fillna(0)

    if hla_class == "I":
        dataframe = dataframe[
            dataframe["Specificity"].astype(str).str.startswith(("A*", "B*"), na=False)
        ]
    elif hla_class == "II":
        dataframe = dataframe[
            dataframe["Specificity"].astype(str).str.startswith("DRB1*", na=False)
        ]

    dataframe = dataframe[dataframe["Raw Value"] >= max(0, int(min_titer))]
    if dataframe.empty:
        return pd.DataFrame(columns=["Локус", "Значение", "MFI"])

    split_parts = dataframe["Specificity"].apply(split_specificity_parts)
    dataframe["Локус"] = split_parts.str[0]
    dataframe["Значение"] = split_parts.str[1]

    dataframe = dataframe.groupby(["Локус", "Значение"], as_index=False).agg(
        {"Raw Value": "mean"}
    )
    dataframe = dataframe.rename(columns={"Raw Value": "MFI"})
    dataframe["MFI"] = dataframe["MFI"].round().astype(int)

    return dataframe.sort_values(
        by=["Локус", "Значение"],
        ascending=[True, True],
    )


def build_combined_titer_dataframe(
    parsed_items: list[ParsedLuminexCsv],
    min_titer: int = 0,
) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []

    class_order = {"I": 0, "II": 1}
    sorted_items = sorted(
        parsed_items,
        key=lambda item: class_order.get(item.hla_class, 99),
    )

    for parsed in sorted_items:
        if parsed.hla_class not in {"I", "II"}:
            raise ValueError("Не удалось определить Class (I/II) в CSV")

        dataframe = build_titer_dataframe(
            antibodies=[item.as_report_dict() for item in parsed.antibodies],
            hla_class=parsed.hla_class,
            min_titer=min_titer,
        )

        if dataframe.empty:
            continue

        frames.append(dataframe)

    if not frames:
        return pd.DataFrame(columns=["Локус", "Значение", "MFI"])

    return pd.concat(frames, ignore_index=True)


# --- Быстрые проверки, есть ли строки для построения отчета ---


def has_titer_rows_from_parsed(
    parsed: ParsedLuminexCsv,
    min_titer: int = 0,
) -> tuple[str, bool]:
    if parsed.hla_class not in {"I", "II"}:
        raise ValueError("Не удалось определить Class (I/II) в CSV")

    dataframe = build_titer_dataframe(
        antibodies=[item.as_report_dict() for item in parsed.antibodies],
        hla_class=parsed.hla_class,
        min_titer=min_titer,
    )
    return parsed.hla_class, not dataframe.empty


def has_titer_rows_from_parsed_group(
    parsed_items: list[ParsedLuminexCsv],
    min_titer: int = 0,
) -> bool:
    dataframe = build_combined_titer_dataframe(
        parsed_items,
        min_titer=min_titer,
    )
    return not dataframe.empty


# --- Генерация Excel по одному классу или по объединенной паре ---


def create_avg_titer_excel_from_parsed(
    parsed: ParsedLuminexCsv,
    output_path: Path,
    min_titer: int = 0,
    *,
    test_date: date | None = None,
    num_register_text: str | None = None,
) -> Path:
    """
    Создаёт Excel-файл по уже распарсенному ParsedLuminexCsv
    и возвращает путь к сохранённому файлу.
    """
    if parsed.hla_class not in {"I", "II"}:
        raise ValueError("Не удалось определить Class (I/II) в CSV")

    patient_norm = format_patient_short_name(parsed.patient)
    report_date = (
        format_ddmmyyyy(test_date) if test_date is not None else parsed.batch_date
    )
    dataframe = build_titer_dataframe(
        antibodies=[item.as_report_dict() for item in parsed.antibodies],
        hla_class=parsed.hla_class,
        min_titer=min_titer,
    )

    excel_file = Path(output_path)
    if excel_file.suffix.lower() != ".xlsx":
        excel_file = excel_file.with_suffix(".xlsx")

    excel_file.parent.mkdir(parents=True, exist_ok=True)

    if excel_file.exists() and excel_file.is_file():
        delete_file(excel_file)

    dataframe.to_excel(excel_file, index=False, startrow=6)

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    center = make_center()
    bold = make_bold()

    paint_header_row(sheet, 7)

    sheet.merge_cells("A1:C1")
    sheet["A1"] = f"Class {parsed.hla_class} LSA"
    sheet["A1"].alignment = center
    sheet["A1"].font = bold

    sheet.merge_cells("A2:C2")
    sheet["A2"] = f"Date: {report_date}"
    sheet["A2"].alignment = center
    sheet["A2"].font = bold

    sheet["A4"] = "Patient:"
    sheet["A4"].alignment = center
    sheet["A4"].font = bold
    sheet.merge_cells("B4:C4")
    sheet["B4"] = patient_norm

    if num_register_text:
        sheet["D4"] = num_register_text
        sheet["D4"].alignment = center
        sheet["D4"].font = bold

    sheet["C5"] = "% PRA:"
    sheet["C5"].alignment = center
    sheet["C5"].font = bold
    sheet["D5"] = parsed.pra
    sheet["D5"].alignment = center

    sheet["A6"] = "Antibodies"
    sheet["A6"].font = bold

    auto_width(sheet, 5, sheet.max_row)

    if num_register_text:
        sheet.column_dimensions["D"].width = max(
            sheet.column_dimensions["D"].width or 0,
            len(str(num_register_text)) + 2,
        )

    center_range(sheet, 7, sheet.max_row, 1, sheet.max_column)
    protect_sheet(sheet, "0")

    workbook.save(excel_file)
    return excel_file


def create_avg_titer_excel_from_parsed_group(
    parsed_items: list[ParsedLuminexCsv],
    output_path: Path,
    min_titer: int = 0,
    *,
    test_date: date | None = None,
    num_register_text: str | None = None,
) -> Path:
    """
    Создаёт один общий Excel для пары CSV Class I и Class II,
    если они относятся к одному patient и batch_date.
    """
    if not parsed_items:
        raise ValueError("Не переданы данные для построения общего Excel")

    first = parsed_items[0]
    report_date = (
        format_ddmmyyyy(test_date) if test_date is not None else first.batch_date
    )
    group_key = build_titer_group_key(first)

    if any(build_titer_group_key(item) != group_key for item in parsed_items[1:]):
        raise ValueError("Нельзя создать общий Excel из CSV разных patient/batch_date")

    dataframe = build_combined_titer_dataframe(
        parsed_items,
        min_titer=min_titer,
    )

    patient_norm = format_patient_short_name(first.patient)

    parsed_class_i = next(
        (item for item in parsed_items if item.hla_class == "I"), None
    )
    parsed_class_ii = next(
        (item for item in parsed_items if item.hla_class == "II"), None
    )

    excel_file = Path(output_path)
    if excel_file.suffix.lower() != ".xlsx":
        excel_file = excel_file.with_suffix(".xlsx")

    excel_file.parent.mkdir(parents=True, exist_ok=True)

    if excel_file.exists() and excel_file.is_file():
        delete_file(excel_file)

    # Заголовок таблицы будет на строке 7
    dataframe.to_excel(excel_file, index=False, startrow=6)

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    center = make_center()
    bold = make_bold()

    paint_header_row(sheet, 7)

    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Class I, II LSA"
    sheet["A1"].alignment = center
    sheet["A1"].font = bold

    sheet.merge_cells("A2:C2")
    sheet["A2"] = f"Date: {report_date}"
    sheet["A2"].alignment = center
    sheet["A2"].font = bold

    sheet["A4"] = "Patient:"
    sheet["A4"].alignment = center
    sheet["A4"].font = bold
    sheet.merge_cells("B4:C4")
    sheet["B4"] = patient_norm

    if num_register_text:
        sheet["D4"] = num_register_text
        sheet["D4"].alignment = center
        sheet["D4"].font = bold

    sheet["C5"] = "% PRA:"
    sheet["C5"].alignment = center
    sheet["C5"].font = bold
    sheet["D5"] = (
        f"{parsed_class_i.pra if parsed_class_i is not None else ''}, {parsed_class_ii.pra if parsed_class_ii is not None else ''}"
    )
    sheet["D5"].alignment = center

    sheet["A6"] = "Antibodies"
    sheet["A6"].font = bold

    auto_width(sheet, 5, sheet.max_row)

    if num_register_text:
        sheet.column_dimensions["D"].width = max(
            sheet.column_dimensions["D"].width or 0,
            len(str(num_register_text)) + 2,
        )

    center_range(sheet, 7, sheet.max_row, 1, sheet.max_column)
    protect_sheet(sheet, "0")

    workbook.save(excel_file)
    return excel_file
