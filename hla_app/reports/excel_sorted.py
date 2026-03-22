"""Генерация отсортированного Excel-результата по одному исследованию.

Здесь реализовано построение Excel-файла для Class I или Class II из набора
антител, включая сортировку, оформление заголовков и сохранение результата
рядом с данными исследования. Файл полезно открывать, если вопрос касается
структуры классических Excel-результатов импорта.
"""

from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from hla_app.data.luminex_parser import parse_fixed_luminex_csv
from hla_app.reports.excel_common import (
    auto_width,
    center_range,
    make_bold,
    make_center,
    paint_header_row,
    protect_sheet,
)
from hla_app.storage.fs_ops import delete_file, replace_file
from hla_app.utils.patient_name import format_patient_short_name

# --- Публичная генерация стандартного Excel-результата исследования ---


def create_sorted_excel_result(csv_file: Path, output_path: Path) -> Path:
    hla_class, batch_date, patient, pra, antibodies = parse_fixed_luminex_csv(csv_file)

    patient_norm = format_patient_short_name(patient)

    if not antibodies:
        df = pd.DataFrame(
            [["None", "----", "----", "----"]],
            columns=["Specificity", "% Positive", "Raw Value", "MFI/LRA"],
        )
    else:
        df = pd.DataFrame(antibodies)

    if antibodies and not df.empty and "Specificity" in df.columns:
        df["Raw Value"] = pd.to_numeric(df.get("Raw Value"), errors="coerce")
        df["sort_key"] = df["Specificity"].astype(str).str.rsplit(":", n=1).str[0]
        df = df.sort_values(
            by=["sort_key", "Raw Value"],
            ascending=[True, False],
            na_position="last",
        )
        df = df.drop(columns="sort_key", errors="ignore")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    tmp_path = output_path.parent / f".tmp__{uuid4().hex}__{output_path.name}"
    book = None

    try:
        if tmp_path.exists() and tmp_path.is_file():
            delete_file(tmp_path)

        df.to_excel(tmp_path, index=False, startrow=6)

        book = load_workbook(tmp_path)
        sheet = book.active

        center = make_center()
        bold = make_bold()

        paint_header_row(sheet, 7)

        sheet.merge_cells("B1:C1")
        sheet["B1"] = f"Class {hla_class} LSA"
        sheet["B1"].alignment = center
        sheet["B1"].font = bold

        sheet.merge_cells("B2:C2")
        sheet["B2"] = f"Date: {batch_date}"
        sheet["B2"].alignment = center
        sheet["B2"].font = bold

        sheet["A4"] = "Patient:"
        sheet["A4"].alignment = center
        sheet["A4"].font = bold
        sheet.merge_cells("B4:C4")
        sheet["B4"] = patient_norm

        sheet["D5"] = "% PRA:"
        sheet["D5"].alignment = center
        sheet["D5"].font = bold
        sheet["E5"] = pra
        sheet["E5"].alignment = center

        sheet["A6"] = "Antibodies"
        sheet["A6"].font = bold

        auto_width(sheet, 5, sheet.max_row)
        center_range(sheet, 7, sheet.max_row, 1, sheet.max_column)
        protect_sheet(sheet, "0")

        book.save(tmp_path)
        book.close()
        book = None

        replace_file(tmp_path, output_path)

        return output_path

    except Exception:
        if book is not None:
            try:
                book.close()
            except Exception:
                pass

        try:
            if tmp_path.exists() and tmp_path.is_file():
                delete_file(tmp_path)
        except Exception:
            pass

        raise
