from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from hla_app.data.luminex_parser import ParsedLuminexCsv
from hla_app.reports.excel_common import (
    auto_width,
    center_range,
    make_bold,
    make_center,
    paint_header_row,
    protect_sheet,
)
from hla_app.storage.fs_ops import delete_file
from hla_app.utils.patient_name import format_patient_short_name
from hla_app.utils.validators import cap_hyphenated_lastname

_INVALID_FILENAME_CHARS_RE = re.compile(r'[\\/:*?"<>|]')


def build_patient_prefix(patient: str) -> str:
    parts = (patient or "").strip().split()
    if not parts:
        return "Пациент"

    surname_raw = re.sub(r"[^A-Za-zА-Яа-яЁё-]+", "", parts[0]).strip("-")
    surname = cap_hyphenated_lastname(surname_raw) if surname_raw else "Пациент"

    rest = "".join(parts[1:])
    letters = re.findall(r"[A-Za-zА-Яа-яЁё]", rest)[:2]
    initials = "".join(letter.upper() for letter in letters)

    prefix = f"{surname}_{initials}" if initials else surname
    return _INVALID_FILENAME_CHARS_RE.sub("_", prefix)


def build_titer_excel_filename(patient: str, hla_class: str) -> str:
    return f"{build_patient_prefix(patient)}_AvgValue_Class_{hla_class}.xlsx"


def build_titer_dataframe(
    *,
    antibodies: list[dict[str, str]],
    hla_class: str,
    min_titer: int = 0,
) -> pd.DataFrame:
    dataframe = pd.DataFrame(antibodies)

    if dataframe.empty or "Specificity" not in dataframe.columns:
        return pd.DataFrame(columns=["Specificity", "Avg Raw Value"])

    dataframe["Raw Value"] = pd.to_numeric(
        dataframe.get("Raw Value"),
        errors="coerce",
    ).fillna(0)

    if hla_class == "I":
        dataframe = dataframe[
            dataframe["Specificity"].astype(str).str.startswith(("A*", "B*"), na=False)
        ]
    elif hla_class == "II":
        dataframe = dataframe[
            dataframe["Specificity"].astype(str).str.startswith("DRB1*", na=False)
        ]

    dataframe = dataframe[dataframe["Raw Value"] >= max(0, int(min_titer))]
    if dataframe.empty:
        return pd.DataFrame(columns=["Specificity", "Avg Raw Value"])

    dataframe["Specificity"] = (
        dataframe["Specificity"].astype(str).str.rsplit(":", n=1).str[0]
    )

    dataframe = dataframe.groupby("Specificity", as_index=False).agg(
        {"Raw Value": "mean"}
    )
    dataframe = dataframe.rename(columns={"Raw Value": "Avg Raw Value"})
    dataframe["Avg Raw Value"] = dataframe["Avg Raw Value"].round().astype(int)

    return dataframe.sort_values(
        by=["Specificity", "Avg Raw Value"],
        ascending=[True, False],
    )


def has_titer_rows_from_parsed(
    parsed: ParsedLuminexCsv,
    min_titer: int = 0,
) -> tuple[str, bool]:
    if parsed.hla_class not in {"I", "II"}:
        raise ValueError("Не удалось определить Class (I/II) в CSV")

    dataframe = build_titer_dataframe(
        antibodies=[item.as_report_dict() for item in parsed.antibodies],
        hla_class=parsed.hla_class,
        min_titer=min_titer,
    )
    return parsed.hla_class, not dataframe.empty


def create_titer_a_b_drb1_excel_from_parsed(
    parsed: ParsedLuminexCsv,
    output_path: Path,
    min_titer: int = 0,
) -> Path:
    """
    Создаёт Excel-файл по уже распарсенному ParsedLuminexCsv
    и возвращает путь к сохранённому файлу.
    """
    if parsed.hla_class not in {"I", "II"}:
        raise ValueError("Не удалось определить Class (I/II) в CSV")

    patient_norm = format_patient_short_name(parsed.patient)
    dataframe = build_titer_dataframe(
        antibodies=[item.as_report_dict() for item in parsed.antibodies],
        hla_class=parsed.hla_class,
        min_titer=min_titer,
    )

    excel_file = Path(output_path)
    if excel_file.suffix.lower() != ".xlsx":
        excel_file = excel_file.with_suffix(".xlsx")

    excel_file.parent.mkdir(parents=True, exist_ok=True)

    if excel_file.exists() and excel_file.is_file():
        delete_file(excel_file)

    dataframe.to_excel(excel_file, index=False, startrow=6)

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    center = make_center()
    bold = make_bold()

    paint_header_row(sheet, 7)

    sheet.merge_cells("A1:B1")
    sheet["A1"] = f"Class {parsed.hla_class} LSA"
    sheet["A1"].alignment = center
    sheet["A1"].font = bold

    sheet.merge_cells("A2:B2")
    sheet["A2"] = f"Date: {parsed.batch_date}"
    sheet["A2"].alignment = center
    sheet["A2"].font = bold

    sheet.merge_cells("A4:B4")
    sheet["A4"] = patient_norm

    sheet["B5"] = "% PRA:"
    sheet["B5"].alignment = center
    sheet["B5"].font = bold
    sheet["C5"] = parsed.pra
    sheet["C5"].alignment = center

    sheet["A6"] = "Antibodies"
    sheet["A6"].font = bold

    auto_width(sheet, 5, sheet.max_row)
    center_range(sheet, 7, sheet.max_row, 1, sheet.max_column)
    protect_sheet(sheet, "0")

    workbook.save(excel_file)
    return excel_file
