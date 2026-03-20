from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def make_center() -> Alignment:
    return Alignment(horizontal="center")


def make_bold() -> Font:
    return Font(bold=True)


def make_gray_fill() -> PatternFill:
    return PatternFill(
        start_color="C0C0C0",
        end_color="C0C0C0",
        fill_type="solid",
    )


def paint_header_row(sheet, header_row: int) -> None:
    fill = make_gray_fill()
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=header_row, column=col).fill = fill


def auto_width(sheet, start_row: int, end_row: int) -> None:
    for col in range(1, sheet.max_column + 1):
        max_width = 0
        for row in range(start_row, end_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                max_width = max(max_width, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(col)].width = max_width + 2


def center_range(sheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    center = make_center()
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.alignment = center


def protect_sheet(sheet, password: str = "0") -> None:
    sheet.protection.set_password(password)
    sheet.protection.enable()
